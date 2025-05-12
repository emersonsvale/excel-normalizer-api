from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.responses import JSONResponse
import openpyxl
import re
from typing import List, Dict, Any
import io
import logging
import traceback
import gc
import os

# Configuração do logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constantes
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB em bytes
BATCH_SIZE = 10000  # Aumentado para 10k linhas por lote
CHUNK_SIZE = 2 * 1024 * 1024  # 2MB chunks para leitura

app = FastAPI(
    title="Excel to JSON Converter",
    description="API para converter arquivos Excel em formato JSON",
    version="1.0.0"
)

def normalize_column_name(column: str) -> str:
    """Normaliza o nome da coluna removendo espaços e convertendo para minúsculas."""
    return re.sub(r'\s+', '', column.lower())

def process_excel_data(workbook: openpyxl.Workbook) -> List[Dict[str, Any]]:
    """Processa os dados da planilha e retorna uma lista de dicionários."""
    try:
        # Pega a primeira planilha
        sheet = workbook.active
        
        # Pega os cabeçalhos
        headers = [normalize_column_name(cell.value) for cell in sheet[1]]
        logger.info(f"Colunas normalizadas: {headers}")
        
        # Verifica se a coluna 'location' existe
        if 'location' not in headers:
            raise HTTPException(
                status_code=400,
                detail=f"Coluna 'Location' não encontrada no arquivo Excel. Colunas disponíveis: {headers}"
            )
        
        # Processa as linhas em lotes
        result = []
        current_batch = []
        total_rows = 0
        processed_rows = 0
        
        # Conta o número total de linhas
        for _ in sheet.iter_rows(min_row=2):
            total_rows += 1
        
        logger.info(f"Total de linhas a processar: {total_rows}")
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            processed_rows += 1
            
            if processed_rows % 100000 == 0:
                logger.info(f"Processando linha {processed_rows} de {total_rows}")
            
            if not row[headers.index('location')]:  # Pula linhas sem location
                continue
                
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = value
            current_batch.append(row_dict)
            
            # Quando o lote estiver cheio, adiciona ao resultado e limpa a memória
            if len(current_batch) >= BATCH_SIZE:
                result.extend(current_batch)
                current_batch = []
                gc.collect()  # Força coleta de lixo
        
        # Adiciona o último lote se houver
        if current_batch:
            result.extend(current_batch)
        
        logger.info(f"Número de registros processados: {len(result)}")
        return result
    
    except Exception as e:
        logger.error(f"Erro ao processar dados: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erro ao processar dados: {str(e)}")

@app.post("/upload")
async def upload_excel(file: UploadFile):
    """
    Endpoint para upload de arquivo Excel.
    
    Args:
        file (UploadFile): Arquivo Excel (.xlsx)
        
    Returns:
        JSONResponse: Dados da planilha em formato JSON
    """
    try:
        logger.info(f"Recebendo arquivo: {file.filename}")
        
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx são aceitos")
        
        # Lê o conteúdo do arquivo em chunks
        contents = bytearray()
        total_size = 0
        
        while chunk := await file.read(CHUNK_SIZE):
            total_size += len(chunk)
            if total_size > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"Arquivo muito grande. Tamanho máximo permitido: {MAX_FILE_SIZE/1024/1024}MB"
                )
            contents.extend(chunk)
            gc.collect()  # Força coleta de lixo após cada chunk
            
        logger.info(f"Tamanho do arquivo: {total_size/1024/1024:.2f}MB")
        
        # Converte para Workbook com otimizações
        workbook = openpyxl.load_workbook(
            io.BytesIO(contents),
            read_only=True,
            data_only=True,  # Carrega apenas valores, não fórmulas
            keep_vba=False   # Não carrega macros
        )
        logger.info(f"Workbook carregado com {len(workbook.sheetnames)} planilhas")
        
        # Processa os dados
        result = process_excel_data(workbook)
        
        # Limpa memória
        del workbook
        del contents
        gc.collect()
        
        return JSONResponse(content=result)
    
    except HTTPException as he:
        logger.error(f"Erro HTTP: {str(he)}")
        raise he
    except Exception as e:
        logger.error(f"Erro inesperado: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

@app.get("/")
async def root():
    """Endpoint raiz com informações sobre a API."""
    return {
        "message": "Bem-vindo à API de conversão Excel para JSON",
        "endpoints": {
            "/upload": "POST - Envie um arquivo Excel para converter em JSON"
        }
    } 